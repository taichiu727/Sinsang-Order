import streamlit as st
import pandas as pd
from datetime import datetime
import os
import io
from PIL import Image
import openpyxl
import tempfile
import base64
from sqlalchemy import create_engine, text
import json
import time
import numpy as np

# Database configuration
DATABASE_URL = "postgresql://neondb_owner:npg_WEzy0a1pQMAl@ep-delicate-forest-a1jhkrwa-pooler.ap-southeast-1.aws.neon.tech/neondb?sslmode=require"

# Initialize database connection
@st.cache_resource
def init_db():
    """Initialize database connection with optimized settings for Neon DB"""
    return create_engine(
        DATABASE_URL,
        connect_args={
            'keepalives': 1,
            'keepalives_idle': 30,
            'keepalives_interval': 10,
            'keepalives_count': 5
        }
    )

def extract_floating_images(excel_file):
    """Extract floating images from Excel file and match them to rows"""
    images_data = []
    
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(excel_file.getvalue())
        tmp_path = tmp_file.name

    try:
        wb = openpyxl.load_workbook(tmp_path)
        ws = wb.active

        for drawing in ws._images:
            row_idx = drawing.anchor._from.row
            img_data = drawing._data()
            try:
                image = Image.open(io.BytesIO(img_data))
                image.thumbnail((300, 300))
                buffered = io.BytesIO()
                image.save(buffered, format="PNG")
                img_str = base64.b64encode(buffered.getvalue()).decode()
                images_data.append({
                    'row_idx': row_idx,
                    'image_data': f'data:image/png;base64,{img_str}'
                })
            except Exception as e:
                st.error(f"Error processing image: {str(e)}")
                continue

        os.unlink(tmp_path)
        return images_data

    except Exception as e:
        st.error(f"Error extracting images: {str(e)}")
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
        return []

def create_tables(engine):
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS orders (
                id SERIAL PRIMARY KEY,
                batch_number INTEGER,
                upload_time TIMESTAMP,
                stall_number TEXT,
                store_name TEXT,
                product_name TEXT,
                color_size TEXT,
                quantity INTEGER,
                unit_price DECIMAL(10, 2),  -- Added unit price column
                received BOOLEAN DEFAULT FALSE,
                missing_quantity INTEGER DEFAULT 0,
                notes TEXT,
                image_data TEXT,
                original_data JSONB,
                parent_order_id INTEGER,
                is_partial_fulfillment BOOLEAN DEFAULT FALSE,
                fulfillment_date TIMESTAMP,
                box_id INTEGER
            )
        """))
        conn.commit()

@st.fragment
def handle_order_fulfillment(engine, row_id, is_received, missing_qty, df):
    """Handle order fulfillment including partial fulfillment logic"""
    try:
        with engine.connect() as conn:
            trans = conn.begin()  # Start transaction
            try:
                # Get the original row from DataFrame
                row = df.loc[row_id]
                original_quantity = int(row['數量'])
                st.write(f"Debug - Original qty: {original_quantity}, Missing qty: {missing_qty}, Received: {is_received}")

                if is_received and missing_qty > 0 and missing_qty < original_quantity:
                    st.write("Debug - Processing partial fulfillment")
                    # Calculate fulfilled quantity
                    fulfilled_qty = original_quantity - missing_qty

                    # Update original order with fulfilled portion
                    update_result = conn.execute(
                        text("""
                            UPDATE orders 
                            SET received = TRUE,
                                quantity = :fulfilled_qty,
                                missing_quantity = 0,
                                fulfillment_date = NOW(),
                                is_partial_fulfillment = TRUE
                            WHERE id = :order_id
                            RETURNING id
                        """),
                        {
                            "fulfilled_qty": fulfilled_qty,
                            "order_id": int(row_id)
                        }
                    )
                    st.write(f"Debug - Updated original order, rows affected: {update_result.rowcount}")

                    # Insert new row for unfulfilled portion
                    insert_result = conn.execute(
                        text("""
                            INSERT INTO orders (
                                batch_number, upload_time, stall_number, store_name,
                                product_name, color_size, quantity, unit_price, received,
                                missing_quantity, notes, image_data, original_data,
                                parent_order_id, is_partial_fulfillment
                            )
                            SELECT 
                                batch_number, NOW(), stall_number, store_name,
                                product_name, color_size, :missing_qty, unit_price, FALSE,
                                :missing_qty, notes, image_data, original_data,
                                id, TRUE
                            FROM orders
                            WHERE id = :order_id
                            RETURNING id
                        """),
                        {
                            "missing_qty": missing_qty,
                            "order_id": int(row_id)
                        }
                    )
                    new_id = insert_result.scalar()
                    st.write(f"Debug - Created new order with ID: {new_id}")

                else:
                    st.write("Debug - Processing regular update")
                    # Regular update for complete fulfillment or no fulfillment
                    update_result = conn.execute(
                        text("""
                            UPDATE orders 
                            SET received = :received,
                                missing_quantity = :missing_qty,
                                fulfillment_date = CASE WHEN :received THEN NOW() ELSE NULL END
                            WHERE id = :order_id
                            RETURNING id
                        """),
                        {
                            "received": is_received,
                            "missing_qty": missing_qty,
                            "order_id": int(row_id)
                        }
                    )
                    st.write(f"Debug - Updated order, rows affected: {update_result.rowcount}")

                trans.commit()
                return True
            except Exception as e:
                trans.rollback()
                st.error(f"Transaction error: {str(e)}")
                return False
    except Exception as e:
        st.error(f"Database connection error: {str(e)}")
        return False

@st.fragment
def save_to_db(engine, df):
    def json_serializable_converter(obj):
        """Convert non-serializable objects to JSON-friendly format"""
        if isinstance(obj, pd.Timestamp):
            return obj.strftime('%Y-%m-%d %H:%M:%S')
        elif isinstance(obj, (np.integer, np.floating)):
            return int(obj) if isinstance(obj, np.integer) else float(obj)
        return obj

    def chunked_execute(conn, data_list, chunk_size=100):
        """Execute inserts in chunks to avoid memory issues"""
        for i in range(0, len(data_list), chunk_size):
            chunk = data_list[i:i + chunk_size]
            try:
                conn.execute(
                    text("""
                        INSERT INTO orders (
                            batch_number, upload_time, stall_number, store_name, 
                            product_name, color_size, quantity, unit_price, received, 
                            missing_quantity, notes, image_data, original_data
                        ) VALUES (
                            :batch_number, :upload_time, :stall_number, :store_name,
                            :product_name, :color_size, :quantity, :unit_price, :received,
                            :missing_quantity, :notes, :image_data, :original_data
                        )
                    """),
                    chunk
                )
                conn.commit()
            except Exception as e:
                st.error(f"Error saving chunk starting at index {i}: {str(e)}")
                raise

    with engine.connect() as conn:
        # Prepare all data first
        data_list = []
        for _, row in df.iterrows():
            row_dict = row.copy()
            
            # Handle different possible column names for upload time
            # Set upload time to current time if not present
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Convert to JSON-serializable dictionary
            original_data = {k: json_serializable_converter(v) for k, v in row_dict.to_dict().items()}
            
            # Compress image data if it's too large
            image_data = row.get('照片', '')
            if len(image_data) > 1000000:  # If larger than 1MB
                try:
                    # Extract base64 data
                    header, base64_data = image_data.split(',', 1)
                    # Decode base64
                    image_bytes = base64.b64decode(base64_data)
                    # Open as PIL Image
                    img = Image.open(io.BytesIO(image_bytes))
                    # Resize if too large
                    max_size = (800, 800)
                    img.thumbnail(max_size, Image.Resampling.LANCZOS)
                    # Save back to base64
                    buffered = io.BytesIO()
                    img.save(buffered, format="PNG", optimize=True, quality=85)
                    compressed_data = base64.b64encode(buffered.getvalue()).decode()
                    image_data = f"{header},{compressed_data}"
                except Exception as e:
                    st.warning(f"Could not compress image: {str(e)}")

            data_list.append({
                "batch_number": row.get('批次號碼/배치 번호'),
                "upload_time": pd.to_datetime(row_dict.get('上傳時間/업로드 시间', datetime.now())),
                "stall_number": row.get('檔口'),
                "store_name": row.get('店名'),
                "product_name": row.get('品名'),
                "color_size": row.get('顏色/尺寸'),
                "quantity": row.get('數量'),
                "unit_price": row.get('單價', 0),
                "received": row.get('到貨/입고', False),
                "missing_quantity": row.get('缺貨數量/부족 수량', 0),
                "notes": row.get('備註', ''),
                "image_data": image_data,
                "original_data": json.dumps(original_data, ensure_ascii=False)
            })

        # Execute in chunks
        try:
            chunked_execute(conn, data_list)
            st.success("數據保存成功 / 데이터 저장 성공")
        except Exception as e:
            st.error(f"保存數據時出錯 / 데이터 저장 오류: {str(e)}")
            conn.rollback()

@st.fragment
def load_from_db(engine):
    with engine.connect() as conn:
        result = conn.execute(text("""
            SELECT 
                id,
                batch_number as "批次號碼/배치 번호",
                upload_time as "上傳時間/업로드 時間",
                stall_number as "檔口",
                store_name as "店名",
                product_name as "品名",
                color_size as "顏色/尺寸",
                quantity as "數量",
                unit_price as "單價",
                received as "到貨/입고",
                missing_quantity as "缺貨數量/부족 수량",
                notes as "備註",
                image_data as "照片",
                box_id
            FROM orders 
            ORDER BY batch_number, upload_time
        """))
        
        df = pd.DataFrame(result.fetchall())
        
        if not df.empty:
            # Set the id as index
            df.set_index('id', inplace=True)
            
            df['上傳時間/업로드 時間'] = pd.to_datetime(df['上傳時間/업로드 時間'])
            
            # Define visible columns
            column_order = [
                '批次號碼/배치 번호',
                '上傳時間/업로드 時間',
                '檔口',
                '店名',
                '品名',
                '顏色/尺寸',
                '數量',
                '單價',  # Added unit price to visible columns
                '到貨/입고',
                '缺貨數量/부족 수량',
                '備註',
                '照片'
            ]
            
            visible_df = df[column_order]
            
            # Store full data in attributes
            visible_df.attrs['full_data'] = df
            
            return visible_df
            
        return df

@st.fragment
def batch_update_db(engine, changes_dict, base_df):
    """Perform batch updates to the database efficiently"""
    try:
        with engine.connect() as conn:
            trans = conn.begin()
            try:
                update_stmt = text("""
                    UPDATE orders 
                    SET received = :received,
                        missing_quantity = :missing_quantity,
                        notes = :notes
                    WHERE batch_number = :batch_number 
                    AND stall_number = :stall_number
                    AND store_name = :store_name
                    AND product_name = :product_name
                    AND color_size = :color_size
                    AND quantity = :quantity
                """)
                
                for row_idx, changes in changes_dict.items():
                    row_idx = int(row_idx)
                    row = base_df.iloc[row_idx]
                    
                    params = {
                        "batch_number": int(row['批次號碼/배치 번호']),
                        "stall_number": str(row['檔口']),
                        "store_name": str(row['店名']),
                        "product_name": str(row['品名']),
                        "color_size": str(row['顏色/尺寸']),
                        "quantity": int(row['數量']),
                        "received": bool(changes.get('到貨/입고', row['到貨/입고'])),
                        "missing_quantity": int(changes.get('缺貨數量/부족 수량', row['缺貨數量/부족 수량']) or 0),
                        "notes": str(changes.get('備註', row['備註']) or '')
                    }
                    
                    conn.execute(update_stmt, params)
                
                trans.commit()
                return True
                
            except Exception as e:
                trans.rollback()
                st.error(f"Transaction error: {str(e)}")
                return False
                
    except Exception as e:
        st.error(f"Database error: {str(e)}")
        return False

def initialize_session_state(engine):
    if "viewport_height" not in st.session_state:
        st.session_state.viewport_height = 800
    if "state" not in st.session_state:
        st.session_state.state = {
            "all_orders_df": load_from_db(engine),
            "latest_upload_df": None,
            "boxes": [],
            "edited_df": None
        }
    if "orders_need_refresh" not in st.session_state:
        st.session_state.orders_need_refresh = True
    if "last_edited_df" not in st.session_state:
        st.session_state.last_edited_df = None
    if "pending_changes" not in st.session_state:
        st.session_state.pending_changes = {}
    

def handle_data_editor_changes(edited_df, engine):
    """Enhanced change handler with partial fulfillment support"""
    if st.session_state.last_edited_df is not None:
        changes = []
        for idx, row in edited_df.iterrows():
            last_row = st.session_state.last_edited_df.iloc[idx]
            
            edit_columns = ['到貨/입고', '缺貨數量/부족 수량', '備註']
            if any(row[col] != last_row[col] for col in edit_columns):
                # Get full data from the original DataFrame
                full_data = st.session_state.last_edited_df.attrs.get('full_data')
                if full_data is not None:
                    row_id = full_data.iloc[idx]['id']
                    
                    # Handle fulfillment with the new logic
                    if handle_order_fulfillment(
                        engine,
                        row_id,
                        bool(row['到貨/입고']),
                        int(row['缺貨數量/부족 수량']) if pd.notna(row['缺貨數量/부족 수량']) else 0,
                        edited_df
                    ):
                        changes.append(idx)

        if changes:
            st.session_state.last_edited_df = edited_df.copy()
            st.toast("✅ Changes saved!")
            st.session_state.orders_need_refresh = True
            return load_from_db(engine)
    
    st.session_state.last_edited_df = edited_df.copy()
    return edited_df

@st.fragment
def orders_editor_section(engine):
    column_config = {
        "照片": st.column_config.ImageColumn("照片", help="商品圖片", width="medium"),
        "單價": st.column_config.NumberColumn("單價", help="單價", format="%.2f", step=0.01),
        "到貨/입고": st.column_config.CheckboxColumn("到貨/입고", help="收貨確認"),
        "缺貨數量/부족 수량": st.column_config.NumberColumn("缺貨數量/부족 수량", help="缺貨數量", min_value=0),
        "備註": st.column_config.TextColumn("備註", help="備註", width="large")
    }


    def handle_edit():
        if 'edited_rows' not in st.session_state.orders_editor:
            return

        edited_rows = st.session_state.orders_editor['edited_rows']
        if not edited_rows:
            return

        try:
            with engine.connect() as conn:
                trans = conn.begin()
                try:
                    update_stmt = text("""
                        UPDATE orders 
                        SET received = :received,
                            missing_quantity = :missing_quantity,
                            notes = :notes
                        WHERE batch_number = :batch_number 
                        AND stall_number = :stall_number
                        AND store_name = :store_name
                        AND product_name = :product_name
                        AND color_size = :color_size
                        AND quantity = :quantity
                    """)
                    
                    for row_idx, row_changes in edited_rows.items():
                        row_idx = int(row_idx)
                        row = st.session_state.state["all_orders_df"].iloc[row_idx]
                        
                        params = {
                            "batch_number": int(row['批次號碼/배치 번호']),
                            "stall_number": str(row['檔口']),
                            "store_name": str(row['店名']),
                            "product_name": str(row['品名']),
                            "color_size": str(row['顏色/尺寸']),
                            "quantity": int(row['數量']),
                            "received": bool(row_changes.get('到貨/입고', row['到貨/입고'])),
                            "missing_quantity": int(row_changes.get('缺貨數量/부족 수량', row['缺貨數量/부족 수량']) or 0),
                            "notes": str(row_changes.get('備註', row['備註']) or '')
                        }
                        
                        conn.execute(update_stmt, params)
                    
                    trans.commit()
                    
                 
                
                except Exception as e:
                    trans.rollback()
                    st.error(f"Error saving changes: {str(e)}")
        
        except Exception as e:
            st.error(f"Database connection error: {str(e)}")

    # Render the data editor with on_change callback
    return st.data_editor(
        st.session_state.state["all_orders_df"],
        column_config=column_config,
        key="orders_editor",
        num_rows="fixed",
        height=st.session_state.viewport_height,
        disabled=[
            "批次號碼/배치 번호", 
            "上傳時間/업로드 시間", 
            "檔口", 
            "店名", 
            "品名", 
            "顏色/尺寸", 
            "數量", 
            "照片"
        ],
        on_change=handle_edit
    )

@st.fragment
def display_order_statistics(df):
    """Display order statistics"""
    if not df.empty:
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            total_orders = len(df)
            received_orders = df['到貨/입고'].sum()
            completion_rate = (received_orders/total_orders*100) if total_orders > 0 else 0
            st.metric("訂單完成率 / 주문 완료율", 
                     f"{completion_rate:.1f}%",
                     f"{received_orders}/{total_orders}")
        
        with col2:
            total_missing = df['缺貨數量/부족 수량'].sum()
            st.metric("總缺貨數量 / 총 부족 수량", f"{total_missing}")
        
        with col3:
            pending_orders = total_orders - received_orders
            st.metric("待處理訂單 / 처리 대기 주문", f"{pending_orders}")
        
        with col4:
            total_batches = df['批次號碼/배치 번호'].nunique()
            st.metric("總批次數 / 총 배치 수", f"{total_batches}")



@st.fragment
def render_upload_tab(engine):
    st.header('上傳訂單 / 주문 업로드')
    uploaded_file = st.file_uploader("選擇Excel文件 / Excel 파일 선택", type=['xlsx'])
    
    if uploaded_file is not None:
        try:
            # Read Excel data
            df = pd.read_excel(uploaded_file)
            
            # Normalize column names
            df.columns = df.columns.str.strip()
            
            # Rename columns to match expected format if needed
            column_mapping = {
                '上傳時間': '上傳時間/업로드 시간',
                '배치번호': '批次號碼/배치 번호',
                '상점명': '店名',
                '품명': '品名',
                '색상/사이즈': '顏色/尺寸',
                '수량': '數量',
                '입고': '到貨/입고',
                '부족수량': '缺貨數量/부족 수량',
                '비고': '備註'
            }
            
            # Apply column mapping
            df.rename(columns=column_mapping, inplace=True)
            
            # Ensure required columns exist
            required_columns = [
                '批次號碼/배치 번호', 
                '上傳時間/업로드 시间', 
                '檔口', 
                '店名', 
                '品名', 
                '顏色/尺寸', 
                '數量'
            ]
            
            for col in required_columns:
                if col not in df.columns:
                    df[col] = ''
            
            # Extract floating images
            images_data = extract_floating_images(uploaded_file)
            
            # Create image HTML for each row
            image_html = {}
            for img in images_data:
                image_html[img['row_idx']] = img["image_data"]
            
            # Add images to DataFrame
            df['照片'] = df.index.map(lambda x: image_html.get(x+1, ''))
            
            # Add status columns if they don't exist
            if '到貨/입고' not in df.columns:
                df['到貨/입고'] = False
            if '缺貨數量/부족 수량' not in df.columns:
                df['缺貨數量/부족 수량'] = 0
            
            # Calculate next batch number
           
            next_batch = 1
            if not st.session_state.state["all_orders_df"].empty:
                next_batch = st.session_state.state["all_orders_df"]['批次號碼/배치 번호'].max() + 1
            
            # Add upload timestamp and batch ID
            df['上傳時間/업로드 시间'] = datetime.now()
            df['批次號碼/배치 번호'] = next_batch
            
            # Store DataFrame for preview
            st.session_state.state["latest_upload_df"] = df
            
            st.success(f'文件上傳成功！找到 {len(images_data)} 張圖片 / 파일이 성공적으로 업로드되었습니다! {len(images_data)}개의 이미지를 찾았습니다')
            
            # Preview the data
            st.subheader("預覽新訂單 / 새 주문 미리보기")
            st.data_editor(
                df,
                column_config={
                    "照片": st.column_config.ImageColumn(
                        "照片",
                        help="商品圖片 / 상품 이미지",
                        width="medium"
                    )
                },
                hide_index=True,
                disabled=True,
                height=st.session_state.viewport_height
            )
            
            # Add confirmation button
            if st.button("確認添加訂單 / 주문 추가 확인"):
                # Save to database
                save_to_db(engine, df)
                
                # Update local DataFrame
                if st.session_state.state["all_orders_df"].empty:
                    st.session_state.state["all_orders_df"] = df.copy()
                else:
                    # Reset index before concatenation
                    existing_df = st.session_state.state["all_orders_df"].reset_index(drop=True)
                    new_df = df.reset_index(drop=True)
                    
                    # Ensure both DataFrames have the same columns
                    columns_to_use = existing_df.columns
                    new_df = new_df[columns_to_use]
                    
                    # Concatenate with reset index
                    st.session_state.state["all_orders_df"] = pd.concat(
                        [existing_df, new_df], 
                        axis=0,
                        ignore_index=True
                    )
                
                st.session_state.state["latest_upload_df"] = None
                st.success("訂單已添加到系統 / 주문이 시스템에 추가되었습니다")
                st.rerun()
                
        except Exception as e:
            st.error(f'錯誤/오류: {str(e)}')

# Tab 2: Track Orders
@st.fragment
def render_tracking_tab():
    """Optimized tracking tab with minimal container creation"""
    if "tracking_tab" not in st.session_state:
        st.session_state.tracking_tab = {
            "container": st.empty(),
            "editor_key": "orders_editor"
        }
    
    with st.session_state.tracking_tab["container"]:
        st.header('追蹤訂單 / 주문 추적')
        
        if st.session_state.state["all_orders_df"].empty:
            st.info('尚無訂單資料 / 주문 데이터가 없습니다')
        else:
            orders_editor_section()

# Tab 3: Shipping Management
@st.fragment
def render_shipping_tab():
    st.header('出貨管理 / 배송 관리')
    

    # Add box deletion controls at the top
    with st.expander("刪除箱子 / 박스 삭제", expanded=False):
        st.warning("請注意：刪除箱子後無法恢復 / 주의: 박스를 삭제하면 복구할 수 없습니다")
        
        if st.button("刪除所有箱子 / 모든 박스 삭제"):
            if delete_all_boxes(init_db()):
                st.session_state.state["all_orders_df"] = load_from_db(init_db())
                
    # Add shipping rate setting at the top
    shipping_rate = st.number_input(
        '運費單價 / 운송비 단가 (NTD/kg)', 
        min_value=0,
        value=4700,
        step=100,
        help="每公斤運費金額 / 킬로그램당 운송비"
    )
    
    # Add export section with export button
    if st.button("匯出裝箱單 / 포장 명세서 내보내기"):
        export_data = generate_combined_box_export(init_db(), shipping_rate)
        if export_data:
            current_date = datetime.now().strftime('%Y%m%d')
            st.download_button(
                label="下載裝箱單 / 포장 명세서 다운로드",
                data=export_data,
                file_name=f"TAI_Export_{current_date}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="export_boxes"
            )
    
    if st.button("刷新數據 / 데이터 새로고침"):
        st.session_state.state["all_orders_df"] = load_from_db(init_db())
        st.rerun()
    
    if not st.session_state.state["all_orders_df"].empty:
        # Get full data with IDs
        full_data = st.session_state.state["all_orders_df"].attrs.get('full_data')
        
        if full_data is not None:
            # Filter for received items that aren't boxed yet
            received_items = full_data[
                (full_data['到貨/입고'] == True) & 
                (full_data['box_id'].isna())
            ].copy()
            
            if not received_items.empty:
                st.subheader("可裝箱商品 / 포장 가능 상품")
                
                # Add selection column and reset index to maintain mapping
                received_items = received_items.reset_index()
                received_items['選擇'] = False
                
                column_config = {
                    "選擇": st.column_config.CheckboxColumn("選擇", help="選擇要裝箱的商品"),
                    "照片": st.column_config.ImageColumn("商品圖片", help="商品圖片", width="medium"),
                    "批次號碼/배치 번호": st.column_config.NumberColumn("批次號碼/배치 번호", help="批次號碼"),
                    "數量": st.column_config.NumberColumn("數量", help="商品數量")
                }
                
                # Display available items with images
                edited_items = st.data_editor(
                    received_items[[
                        '選擇', 'id', '批次號碼/배치 번호', '檔口', '店名', 
                        '品名', '顏色/尺寸', '數量', '照片'
                    ]],
                    column_config=column_config,
                    use_container_width=True,
                    key="available_items",
                    hide_index=True
                )
                
                # Handle boxing
                selected_items = edited_items[edited_items['選擇']]
                if not selected_items.empty:
                    box_weight = st.number_input(
                        '箱子重量 / 박스 무게 (kg)', 
                        min_value=0.0,
                        step=0.1
                    )
                    
                    if st.button("創建箱子 / 박스 만들기"):
                        try:
                            with init_db().connect() as conn:
                                trans = conn.begin()
                                try:
                                    # Create new box
                                    result = conn.execute(
                                        text("""
                                            INSERT INTO boxes (
                                                created_at, 
                                                weight
                                            ) VALUES (
                                                NOW(), 
                                                :weight
                                            ) RETURNING id
                                        """),
                                        {
                                            "weight": float(box_weight)
                                        }
                                    )
                                    box_id = result.scalar()
                                    
                                    # Get order IDs directly from selected items
                                    order_ids = selected_items['id'].tolist()
                                    
                                    # Update each selected order with the box_id
                                    for order_id in order_ids:
                                        conn.execute(
                                            text("""
                                                UPDATE orders 
                                                SET box_id = :box_id 
                                                WHERE id = :order_id
                                            """),
                                            {
                                                "box_id": int(box_id),
                                                "order_id": int(order_id)
                                            }
                                        )
                                    
                                    trans.commit()
                                    st.toast("✅ 箱子已創建 / 박스가 생성되었습니다")
                                    st.session_state.state["all_orders_df"] = load_from_db(init_db())
                                    st.rerun()
                                    
                                except Exception as e:
                                    trans.rollback()
                                    st.error(f"Error in transaction: {str(e)}")
                                    
                        except Exception as e:
                            st.error(f"Database connection error: {str(e)}")
            else:
                st.info("沒有可裝箱的商品 / 포장할 수 있는 상품이 없습니다")
            
            # Display existing boxes
            st.subheader("已裝箱商品 / 포장된 상품")
            
            # Query to get boxes with their items
            boxes_query = text("""
                SELECT 
                    b.id as box_id,
                    b.created_at,
                    b.weight,
                    o.id as order_id,
                    o.batch_number,
                    o.stall_number,
                    o.store_name,
                    o.product_name,
                    o.color_size,
                    o.quantity,
                    o.unit_price,
                    o.image_data
                FROM boxes b
                LEFT JOIN orders o ON o.box_id = b.id
                ORDER BY b.created_at DESC, o.id
            """)
            
            with init_db().connect() as conn:
                boxes_result = conn.execute(boxes_query)
                boxes_data = pd.DataFrame(boxes_result.fetchall())
                
                if not boxes_data.empty:
                    # Calculate totals for each box
                    boxes_data['item_total'] = boxes_data['quantity'] * boxes_data['unit_price']
                    
                    # Group by box
                    for box_id, box_group in boxes_data.groupby('box_id'):
                        box_info = box_group.iloc[0]
                        box_letter = chr(65 + int(box_id) - 1)  # Convert to letter (A=1, B=2, etc.)
                        box_total = box_group['item_total'].sum()
                        box_shipping = box_info['weight'] * shipping_rate
                        
                        with st.expander(
                            f"箱子 {box_letter} - "
                            f"{box_info['created_at'].strftime('%Y-%m-%d %H:%M')} "
                            f"(重量: {box_info['weight']}kg, 運費: ${box_shipping:,.0f})"
                        ):
                            # Display box items with unit prices and totals
                            st.dataframe(
                                box_group[[
                                    'batch_number', 'stall_number', 'store_name',
                                    'product_name', 'color_size', 'quantity', 
                                    'unit_price', 'item_total', 'image_data'
                                ]].assign(**{
                                    'batch_number': box_group['batch_number'].astype(int),
                                    'quantity': box_group['quantity'].astype(int),
                                    'unit_price': box_group['unit_price'].map('${:,.0f}'.format),
                                    'item_total': box_group['item_total'].map('${:,.0f}'.format)
                                }),
                                column_config={
                                    "image_data": st.column_config.ImageColumn(
                                        "商品圖片",
                                        help="商品圖片",
                                        width="medium"
                                    ),
                                    "batch_number": "批次號碼/배치 번호",
                                    "stall_number": "檔口",
                                    "store_name": "店名",
                                    "product_name": "品名",
                                    "color_size": "顏色/尺寸",
                                    "quantity": "數量",
                                    "unit_price": "單價",
                                    "item_total": "總價"
                                },
                                use_container_width=True,
                                hide_index=True
                            )
                            
                            # Add delete button for individual box
                            col1, col2 = st.columns([3, 1])
                            with col2:
                                if st.button(f"刪除箱子 {box_letter} / 박스 삭제", key=f"delete_box_{box_id}"):
                                    if delete_box(init_db(), box_id):
                                        st.session_state.state["all_orders_df"] = load_from_db(init_db())
                else:
                    st.info("沒有已裝箱的商品 / 포장된 상품이 없습니다")
    else:
        st.info('請先上傳訂單文件 / 주문 파일을 먼저 업로드하세요')
        
def create_boxes_table(engine):
    """Create boxes table if it doesn't exist"""
    with engine.connect() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS boxes (
                id SERIAL PRIMARY KEY,
                created_at TIMESTAMP NOT NULL DEFAULT NOW(),
                weight DECIMAL(10,2),
                shipping_fee DECIMAL(10,2),
                status TEXT DEFAULT 'pending'
            )
        """))
        conn.commit()

def get_box_contents(engine, box_id):
    """Get all items in a specific box"""
    with engine.connect() as conn:
        result = conn.execute(
            text("""
                SELECT 
                    o.batch_number,
                    o.stall_number,
                    o.store_name,
                    o.product_name,
                    o.color_size,
                    o.quantity,
                    o.fulfillment_date
                FROM orders o
                WHERE o.box_id = :box_id
                ORDER BY o.fulfillment_date
            """),
            {"box_id": box_id}
        )
        return pd.DataFrame(result.fetchall())

def update_box_status(engine, box_id, status):
    """Update box status"""
    with engine.connect() as conn:
        try:
            conn.execute(
                text("UPDATE boxes SET status = :status WHERE id = :box_id"),
                {"status": status, "box_id": box_id}
            )
            conn.commit()
            return True
        except Exception as e:
            st.error(f"Error updating box status: {str(e)}")
            return False

def remove_item_from_box(engine, order_id):
    """Remove a single item from its box"""
    with engine.connect() as conn:
        try:
            conn.execute(
                text("UPDATE orders SET box_id = NULL WHERE id = :order_id"),
                {"order_id": order_id}
            )
            conn.commit()
            return True
        except Exception as e:
            st.error(f"Error removing item from box: {str(e)}")
            return False

@st.fragment
def delete_box(engine, box_id):
    """Delete a box and unlink its orders"""
    try:
        with engine.connect() as conn:
            trans = conn.begin()
            try:
                # First reset the box_id in orders to unlink them
                conn.execute(
                    text("UPDATE orders SET box_id = NULL WHERE box_id = :box_id"),
                    {"box_id": int(box_id)}
                )
                
                # Then delete the box
                conn.execute(
                    text("DELETE FROM boxes WHERE id = :box_id"),
                    {"box_id": int(box_id)}
                )
                
                trans.commit()
                st.success(f"箱子已刪除 / 박스가 삭제되었습니다")
                return True
                
            except Exception as e:
                trans.rollback()
                st.error(f"Error in transaction: {str(e)}")
                return False
                
    except Exception as e:
        st.error(f"Database connection error: {str(e)}")
        return False
@st.fragment
def delete_all_boxes(engine):
    """Delete all boxes and reset all order box assignments"""
    try:
        with engine.connect() as conn:
            trans = conn.begin()
            try:
                # First reset all box_id references in orders
                conn.execute(text("UPDATE orders SET box_id = NULL"))
                
                # Then delete all boxes
                conn.execute(text("DELETE FROM boxes"))
                
                trans.commit()
                st.success("所有箱子已刪除 / 모든 박스가 삭제되었습니다")
                return True
                
            except Exception as e:
                trans.rollback()
                st.error(f"Error in transaction: {str(e)}")
                return False
                
    except Exception as e:
        st.error(f"Database connection error: {str(e)}")
        return False

# Add this to the shipping tab rendering function
@st.fragment
def add_box_deletion_controls(engine):
    """Add box deletion controls to the shipping tab"""
    with st.expander("刪除箱子 / 박스 삭제", expanded=False):
        st.warning("請注意：刪除箱子後無法恢復 / 주의: 박스를 삭제하면 복구할 수 없습니다")
        
        if st.button("刪除所有箱子 / 모든 박스 삭제"):
            if delete_all_boxes(engine):
                st.session_state.state["all_orders_df"] = load_from_db(engine)
             
    
def generate_combined_box_export(engine, shipping_rate=4700):
    """Generate Excel export data for all boxes with combined calculations"""
    import pandas as pd
    from datetime import datetime
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # Create Excel writer
    output = io.BytesIO()
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Export Report"

    # Define styles
    header_font = Font(name='微軟正黑體', bold=True)
    normal_font = Font(name='微軟正黑體')
    header_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    centered_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    right_alignment = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Get box data from database with unit prices
    with engine.connect() as conn:
        boxes_query = text("""
            SELECT 
                b.id as box_id,
                b.created_at,
                b.weight,
                o.batch_number,
                o.stall_number,
                o.store_name,
                o.product_name,
                o.color_size,
                o.quantity,
                o.unit_price
            FROM boxes b
            LEFT JOIN orders o ON o.box_id = b.id
            ORDER BY b.id, o.batch_number
        """)
        result = conn.execute(boxes_query)
        boxes_data = pd.DataFrame(result.fetchall())

        if boxes_data.empty:
            return None

        # Convert Decimal columns to float for calculations
        boxes_data['unit_price'] = boxes_data['unit_price'].astype(float)
        boxes_data['quantity'] = boxes_data['quantity'].astype(float)
        boxes_data['weight'] = boxes_data['weight'].astype(float)
        
        # Calculate totals
        boxes_data['item_total'] = boxes_data['quantity'] * boxes_data['unit_price']
        total_order_fee = float(boxes_data['item_total'].sum())
        service_fee = float(total_order_fee * 0.03)
        total_weight = float(boxes_data.groupby('box_id')['weight'].first().sum())
        shipping_fee = float(total_weight * shipping_rate)
        grand_total = total_order_fee + service_fee + shipping_fee

        # Write company header
        worksheet['B1'] = "CASTLE"
        worksheet['B1'].font = Font(name='微軟正黑體', bold=True, size=14)

        # Write summary section
        summary_data = [
            ['總貨價', f'{total_order_fee:,.0f}', '總重量', f'{total_weight:.1f} kg'],
            ['服務費 (3%)', f'{service_fee:,.0f}', '運費單價', f'{shipping_rate:,.0f} NTD/kg'],
            ['運費總額', f'{shipping_fee:,.0f}', '', ''],
            ['總金額', f'{grand_total:,.0f}', '', '']
        ]

        current_row = 3
        for row_data in summary_data:
            for col_idx, value in enumerate(row_data):
                cell = worksheet[f'{get_column_letter(col_idx + 2)}{current_row}']
                cell.value = value
                cell.font = normal_font
                cell.alignment = left_alignment
            current_row += 1

        # Add space before items table
        current_row += 2

        # Write items table header
        headers = ['箱號', '批次號碼', '檔口', '店名', '品名', '顏色/尺寸', '數量', '單價', '總價']
        for col_idx, header in enumerate(headers):
            col = get_column_letter(col_idx + 2)
            cell = worksheet[f'{col}{current_row}']
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = thin_border

        # Write items data
        current_row += 1
        for _, row in boxes_data.iterrows():
            box_letter = chr(65 + int(row['box_id']) - 1)
            data = [
                box_letter,
                row['batch_number'],
                row['stall_number'],
                row['store_name'],
                row['product_name'],
                row['color_size'],
                row['quantity'],
                row['unit_price'],
                row['item_total']
            ]

            for col_idx, value in enumerate(data):
                col = get_column_letter(col_idx + 2)
                cell = worksheet[f'{col}{current_row}']
                cell.value = value
                cell.font = normal_font
                cell.border = thin_border
                
                if col_idx in [0, 1, 6]:  # box letter, batch number, quantity
                    cell.alignment = centered_alignment
                elif col_idx in [7, 8]:  # prices
                    cell.alignment = right_alignment
                    cell.number_format = '#,##0'
                else:
                    cell.alignment = left_alignment
            
            current_row += 1

        # Set column widths
        column_widths = {
            'B': 8,   # 箱號
            'C': 12,  # 批次號碼
            'D': 10,  # 檔口
            'E': 20,  # 店名
            'F': 30,  # 品名
            'G': 20,  # 顏色/尺寸
            'H': 8,   # 數量
            'I': 12,  # 單價
            'J': 12   # 總價
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

    # Save workbook
    try:
        workbook.save(output)
        output.seek(0)
        return output
    except Exception as e:
        st.error(f"Error saving Excel file: {str(e)}")
        return None

def add_box_export_section():
    """Add export section with box statistics and export button"""
    with init_db().connect() as conn:
        # Get box statistics
        stats_query = text("""
            SELECT 
                COUNT(DISTINCT b.id) as total_boxes,
                SUM(b.weight) as total_weight,
                SUM(b.shipping_fee) as total_shipping,
                COUNT(o.id) as total_items,
                SUM(o.quantity) as total_quantity
            FROM boxes b
            LEFT JOIN orders o ON o.box_id = b.id
            WHERE b.id IS NOT NULL
        """)
        
        result = conn.execute(stats_query)
        stats = result.fetchone()
        
        if stats and stats[0] > 0:
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("總箱數 / 총 박스 수", f"{stats[0]}")
            with col2:
                st.metric("總重量 / 총 무게", f"{stats[1]:.1f} kg")
            with col3:
                st.metric("總運費 / 총 운송비", f"${stats[2]:.2f}")
            
            # Generate single Excel file with all boxes
            box_data = generate_combined_box_export(init_db())
            if box_data:
                current_date = datetime.now().strftime('%Y.%m.%d')
                
                st.download_button(
                    label="下載所有箱子清單 / 모든 박스 목록 다운로드",
                    data=box_data,
                    file_name=f"TAI {current_date}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="export_all_boxes"
                )

def main():
    st.set_page_config(page_title="Order Tracking", layout="wide")
    
    # Initialize database and tables
    engine = init_db()
    create_tables(engine)
    create_boxes_table(engine)
    
    # Initialize session state
    initialize_session_state(engine)
    
    st.title('物流追蹤系統 / 물류 추적 시스템')

    tab1, tab2, tab3 = st.tabs(['上傳訂單/주문 업로드', '追蹤訂單/주문 추적', '出貨管理/배송 관리'])

    with tab1:
        render_upload_tab(engine)  # Pass engine as an argument
    with tab2:
        st.header('追蹤訂單 / 주문 추적')
        
        if st.session_state.state["all_orders_df"].empty:
            st.info('尚無訂單資料 / 주문 데이터가 없습니다')
        else:
            orders_editor_section(engine)
    with tab3:
        render_shipping_tab()

if __name__ == "__main__":
    main()